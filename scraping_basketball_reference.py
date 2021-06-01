import bs4
import openpyxl
import pandas as pd
import requests

readFile = pd.read_excel("NBA_App_v2.xlsx", header=None)
for readLineTeam in range(readFile.shape[0]):
    team = readFile.iloc[readLineTeam][0]
    URL = "https://www.basketball-reference.com/teams/" + team + "/2021.html"
    fileWrite = openpyxl.load_workbook("NBA Teams/" + team + ".xlsx")
    sheetWrite = fileWrite["Sheet1"]

    requests.get(URL)

    webPage = bs4.BeautifulSoup(requests.get(URL, {}).text, "lxml")

    # Lecture Tableau "Advanced" pour déterminer WS
    advancedTable = webPage.find(name="table", attrs={"id": "advanced"})
    listPlayers = []
    wsList = []

    advancedTablePlayers = advancedTable.tbody.children
    # On génère la liste des joueurs à partir de basket-ball reference pour une équipe
    for row in advancedTablePlayers:
        if type(row) == bs4.NavigableString:
            continue
        rowTableAdvanced = row.find("td")
        rowTableAdvanced = str(rowTableAdvanced)
        splitUnPremiereMoitie, splitUnDeuxiemeMoitie = rowTableAdvanced.split('.html">')

        player, waste = splitUnDeuxiemeMoitie.split("</a>")
        listPlayers.append(player)

    # On génère la liste des Win Share des joueurs de l'équipe à partir de basket-ball reference
    ## Le Win Share correspond au nombre de victoires dont le joueur est "responsable". Plus le WS est élevé, meilleur
    ## est le joueur
    for row2 in advancedTable.tbody.children:
        if type(row2) == bs4.NavigableString:
            continue
        row2TableAdvanced = row2.find(name="td", attrs={"data-stat": "ws"})
        row2TableAdvanced = str(row2TableAdvanced)
        waste, keptPart = row2TableAdvanced.split('"ws">')
        ws, waste = keptPart.split("</")
        wsList.append(float(ws))

    print(listPlayers)
    print(wsList)

    # On utilise les deux listes générés pour remplir le fichier Excel de l'équipe
    for i in range(len(listPlayers)):
        sheetWrite.cell(row=i + 2, column=5).value = listPlayers[i]
        sheetWrite.cell(row=i + 2, column=6).value = wsList[i]

    # Lecture Injury Report : Recensement des joueurs blessés/inactifs
    from bs4 import Comment, BeautifulSoup
    for i in range(2, 10):
        sheetWrite.cell(row=i, column=7).value = ""

    playersInjured = []

    # Lorsqu'une équipe ne présente pas de joueurs blessés, cela créé un décalage sur les tableaux de la page de stats.
    # On introduit donc une variable qui sera égale à 1 lorsqu'une équipe ne présente pas de blessé, afin de compenser
    # ce décalage
    adjustement = 0

    injuriesReport = webPage.find_all(text=lambda text: isinstance(text, Comment))[33]
    injuriesReportSoup = BeautifulSoup(injuriesReport, 'lxml')
    injuriesReportSoup = injuriesReportSoup.tbody
    # On teste si le tableau car ce n'est pas garanti !
    if injuriesReportSoup is not None:
        for rowInjuriesReport in injuriesReportSoup:
            if type(rowInjuriesReport) == bs4.NavigableString:
                continue
            rowInjuriesPlayer = rowInjuriesReport.find("th")
            player = rowInjuriesPlayer.a.text
            rowInjuriesNote = rowInjuriesReport.find_all(name="td", attrs={"data-stat": "note"})[0]
            playerProbable = False
            if "PROBABLE" in rowInjuriesNote.text or "probable" in rowInjuriesNote.text:
                playerProbable = True
            if not playerProbable:
                playersInjured.append(player)

        for j in range(len(playersInjured)):
            sheetWrite.cell(row=j + 2, column=7).value = playersInjured[j]
    # Si le tableau des blessés n'est pas présent, on met la variable adjustement à 1 pour pouvoir lire les stats quand même
    else:
        sheetWrite.cell(row=2, column=7).value = "NA"
        adjustement = 1

    # Lecture Matchs Précédents pour déterminer CE
    timelineList = []
    timelines = webPage.find_all(name="ul", attrs={"class": "timeline"})

    for timeline in timelines:
        for row in timeline:
            if type(row) == bs4.NavigableString:
                continue
            row = row.find("span").text
            if row is "":
                event = "Off Day"
            else:
                event = "Game Day"
            timelineList.append(event)

    for k in range(len(timelineList)):
        sheetWrite.cell(row=k + 2, column=9).value = timelineList[k]

    # Lecture Stats pour déterminer les trois paramètres stats de l'équipe, la variable adjustement permet de lire le tableau
    # qu'il y ait des blessés ou non
    # Les paramètres sont calculés en effectuant le ratio entre les stats de l'équipe contre ses adversaires sur les stats des
    # équipes adverses contre eux. Un ratio supérieur à 1 pour un paramètre indique que l'équipe tend à dominer ses adversaires
    # sur ce domaine

    miscTeamStats = webPage.find_all(text=lambda text: isinstance(text, Comment))[39 - adjustement]
    miscTeamStatsSoup = BeautifulSoup(miscTeamStats, 'lxml')
    miscTeamStats = miscTeamStatsSoup.tbody
    miscTeamStatsRow = miscTeamStats.find_all("tr")[0]
    # Paramètre lié à la précision au tir
    eFGTeam = miscTeamStatsRow.find_all(name="td", attrs={"data-stat": "efg_pct"})[0].text
    eFGOpponent = miscTeamStatsRow.find_all(name="td", attrs={"data-stat": "opp_efg_pct"})[0].text
    eFG = float(eFGTeam) / float(eFGOpponent)
    # Paramètre lié aux rebonds
    ORB = miscTeamStatsRow.find_all(name="td", attrs={"data-stat": "orb_pct"})[0].text
    DRB = miscTeamStatsRow.find_all(name="td", attrs={"data-stat": "drb_pct"})[0].text
    RBD = float(ORB) / (100 - float(DRB))
    sheetWrite.cell(row=2, column=1).value = eFG
    sheetWrite.cell(row=2, column=2).value = RBD

    teamOppStats = webPage.find_all(text=lambda text: isinstance(text, Comment))[38 - adjustement]
    teamOppStatsSoup = BeautifulSoup(teamOppStats, 'lxml')
    teamStats = teamOppStatsSoup.find_all("tbody")[0]
    oppStats = teamOppStatsSoup.find_all("tbody")[1]

    # Paramètre lié à la capacité d'une équipe de convertir ses possessions de balle en points
    astTeam = float(teamStats.find_all(name="td", attrs={"data-stat": "ast"})[0].text)
    astOpp = float(oppStats.find_all(name="td", attrs={"data-stat": "opp_ast"})[0].text)
    tovTeam = float(teamStats.find_all(name="td", attrs={"data-stat": "tov"})[0].text)
    tovOpp = float(oppStats.find_all(name="td", attrs={"data-stat": "opp_tov"})[0].text)
    atrTeam = astTeam / tovTeam
    atrOpp = astOpp / tovOpp
    ATR = atrTeam / atrOpp
    sheetWrite.cell(row=2, column=3).value = ATR

    fileSaveName = "NBA Teams/" + team + ".xlsx"
    fileWrite.save(fileSaveName)
