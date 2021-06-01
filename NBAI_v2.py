import pandas as pd
from sklearn.calibration import calibration_curve, CalibratedClassifierCV
from sklearn.ensemble import RandomForestClassifier, VotingClassifier, GradientBoostingClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import cross_val_score
from sklearn.metrics import accuracy_score, brier_score_loss
from sklearn.naive_bayes import GaussianNB
from sklearn.neural_network import MLPClassifier
from sklearn.svm import SVC
import numpy as np
from matplotlib import pyplot
import openpyxl
from decimal import Decimal
from datetime import date
from operator import truediv
import bs4
import requests

X_train_ext = pd.read_excel("xtrain_all_para.xlsx", header=None)
Y_train_ext = pd.read_excel("ytrain_extended.xlsx", header=None)

cut_off = 600
X_train = X_train_ext[:cut_off]
X_test = X_train_ext[cut_off:]
Y_train = Y_train_ext[:cut_off]
Y_test = Y_train_ext[cut_off:]


def calibrate(model, method):
    model = CalibratedClassifierCV(model, method=method, cv=5)
    return model


method = 'sigmoid'

log_clf = LogisticRegression(max_iter=5000)
bdt_clf = GradientBoostingClassifier(n_estimators=150,
                                     max_depth=1,
                                     min_samples_split=2,
                                     min_samples_leaf=2,
                                     max_features='auto',
                                     learning_rate=0.2)
nvb_clf = GaussianNB()
rnd_clf = RandomForestClassifier(n_estimators=75,
                                 max_features='sqrt',
                                 min_samples_split=10,
                                 min_samples_leaf=10,
                                 max_depth=3)
svm_clf = SVC(kernel='rbf', gamma=1, C=10, probability=True)
mlp_clf = MLPClassifier(hidden_layer_sizes=(5, 5),
                        activation='relu',
                        solver='lbfgs',
                        alpha=0.005,
                        learning_rate_init=0.02,
                        max_iter=2000)

#estimators = [('rf', rnd_clf), ('svc', svm_clf), ('nvb', nvb_clf), ('mlp', mlp_clf)]
estimators = [('lr', log_clf), ('bdt', bdt_clf), ('nvb', nvb_clf)]

voting_clf = VotingClassifier(
    estimators=estimators, voting='soft'
)
voting_clf.fit(X_train, np.ravel(Y_train))

cal_clf = []

for clf in (log_clf, bdt_clf, nvb_clf, rnd_clf, svm_clf):
    clf.fit(X_train, np.ravel(Y_train))
    clf = calibrate(clf, method)
    clf.fit(X_train, np.ravel(Y_train))
    y_guess_proba = clf.predict_proba(X_test)
    y_guess = clf.predict(X_test)
    cal_clf.append(clf)
    print(brier_score_loss(Y_test, y_guess_proba[:, 1]))

voting_clf.fit(X_train, np.ravel(Y_train))
y_guess_proba = voting_clf.predict_proba(X_test)
y_guess = voting_clf.predict(X_test)
cal_clf.append(voting_clf)
print(voting_clf, accuracy_score(y_guess, Y_test), brier_score_loss(Y_test, y_guess_proba[:, 1]))

probs = voting_clf.predict_proba(X_test)

# Tracé de la courbe de calibration des probabilités prédites sur le dataset de Test
fop, mpv = calibration_curve(np.ravel(Y_test), probs[:, 1], n_bins=5, normalize=False, strategy='quantile')
pyplot.plot(mpv, fop, marker='.', label='calibrated ' + method)
pyplot.plot([0, 1], [0, 1], linestyle='--')
pyplot.legend()
pyplot.show()

probs_total = voting_clf.predict_proba(X_train_ext)

y_guess_train_ext = voting_clf.predict(X_train_ext)
y_guess_train = voting_clf.predict(X_train)

model = cal_clf[len(cal_clf) - 1]
proba = model.predict_proba(X_train_ext)
proba_home = proba[:, 1]
proba_home = proba_home.reshape((925, 1))

scores = cross_val_score(voting_clf, X_train_ext, np.ravel(Y_train_ext), cv=3)
print(np.mean(scores))

colf = [1]

file = pd.read_excel("NBA_App_v2.xlsx", header=0, converters={col: float for col in colf}, sheet_name=None)
# Lecture des équipes concernées par les matchs du jour et récupération des statistiques d'intérêt
file_stats = file["Sheet1"]

# Formattage des valeurs à 3 chiffres après la virgule
float_formatter = "{:.3f}".format

# Liste regroupant les stats d'intérêts : TS ==> TeamStats
TS = []
number_of_games = file_stats.shape[0] / 2

for readLine in range(file_stats.shape[0]):
    team = file_stats.iloc[readLine]["Team"]

    # Liste regroupant les stats d'équipe
    Team_Stats = []

    # On lit le fichier Excel associé à l'équipe lue
    readFileName = "NBA Teams/" + team + ".xlsx"
    colf = [0, 1, 2, 5]
    readFile = pd.read_excel(readFileName, header=0, converters={col: float for col in colf})
    # Lecture des stats de l'équipe lue
    ## Paramètre lié au secteur du rebond
    Team_Stats.append(readFile.iloc[0]["RBD"])
    ## Paramètre lié à la précision au tir
    Team_Stats.append(readFile.iloc[0]["eFG"])
    ## Paramètre lié à la capacité d'une équipe de convertir une possesion de balles en points (Assist to Turnover Ratio)
    Team_Stats.append(readFile.iloc[0]["ATR"])

    # Variable qui stockera le Win Share cumulé des joueurs blessés
    wsfTotalInjured = 0


    ws_team = readFile["WS"]
    sum_of_ws = np.sum(ws_team)

    for readLineInjured in range(readFile.shape[0]):
        injuredPlayer = readFile.iloc[readLineInjured]["Injured Players"]
        if injuredPlayer is not None:
        # Détermination du Win Share total des blessés de l'équipe lue
            for readLineRoster in range(readFile.shape[0]):
                player = readFile.iloc[readLineRoster]["Players"]
                if injuredPlayer == player:
                    ws = readFile.iloc[readLineRoster]["WS"]
                    wsfTotalInjured = wsfTotalInjured + ws
                elif player is None:
                    break

    wsfTotalRemaining = 1 - (wsfTotalInjured / sum_of_ws)
    print(team, "wsf injured = ", wsfTotalInjured)
    Team_Stats.append(wsfTotalRemaining)
    print(team, "wsf remaining = ", wsfTotalRemaining)
    print(team, "sum ws = ", sum_of_ws)

# Détermination du Calendar Effect
    # Variable qui stockera l'impact du calendrier. Une valeur élevée indique une équipe reposée
    ceTeam = 0

    # On détermine là où on doit s'arrêter dans la "Timeline" de l'équipe lue
    numberGamesPlayed = file_stats.iloc[readLine]["Number of Games Played"]
    countGamesPlayed = 0
    for readLineCalendar in range(readFile.shape[0]):
        event = readFile.iloc[readLineCalendar]["Timeline"]
        if event == "Game Day":
            countGamesPlayed = countGamesPlayed + 1
            if countGamesPlayed == numberGamesPlayed + 1:
                stop = readLineCalendar
                break

    # On détermine si l'équipe lue a un match la veille ou le lendemain (Back To Back)
    eventOneDayAfter = readFile.iloc[stop + 1]["Timeline"]
    eventOneDayBefore = readFile.iloc[stop - 1]["Timeline"]
    backToBack = False
    if eventOneDayAfter == "Game Day" or eventOneDayBefore == "Game Day":
        backToBack = True
        # Application d'un malus en cas de Back To Back --> Plus exigeant physiquement de jouer deux matchs à la suite
        ceTeam = ceTeam - 0.10

    readLineCE = stop
    numberGameDays = 0
    numberDaysWatched = 0
    done = False
    while not done:
        event = readFile.iloc[readLineCE]["Timeline"]
        numberDaysWatched = numberDaysWatched + 1
        if event == "Off Day":
            ceTeam = ceTeam + 0.2
        else:
            numberGameDays = numberGameDays + 1
            # On teste le cas particulier où l'équipe peut être amené 3 matchs en 4 jours (scénario le plus exigeant physiquement)
            if numberGameDays == 3 and numberDaysWatched == 4:
                if backToBack:
                    ceTeam = 0.15
                else:
                    ceTeam = 0.25
                done = True
        if numberDaysWatched == 5:
            done = True
        readLineCE = readLineCE - 1
    Team_Stats.append(ceTeam)
    TS.append(Team_Stats)

# Compteur pour parcourir la liste WSF
k = 0

indiceSecuriteHome = 1.07
indiceSecuriteAway = 1.11

# Ouverture du template pour le rapport du jour
workbook = openpyxl.load_workbook("Today.xlsx")
stw = workbook["Sheet1"]

# Ecriture des paramètres d'entrée pour chaque match, on fait le ratio des stats de l'équipe à domicile sur
# les stats de l'équipe à l'extérieur. Un ratio supérieur à 1 indique que l'équipe à domicile est plus performante
# pour un paramètre donné.

for i in range(1, len(TS), 2):
    teamAway = file_stats.iloc[i - 1][0]
    teamHome = file_stats.iloc[i][0]
    stw.cell(row=k + 2, column=2).value = teamAway + " @ " + teamHome

    stats_final = list(map(truediv, TS[i], TS[i - 1]))
    print(stats_final)
    for j in range(5):
        stw.cell(row=k + 2, column=j + 3).value = stats_final[j]
    stats_final = np.array(stats_final)
    stats_final = stats_final.reshape((1, -1))

    # On utilise le Voting Classifier, qui est le dernier élément de la liste cal_clf
    model = cal_clf[len(cal_clf) - 1]
    proba = model.predict_proba(stats_final)

    np.set_printoptions(formatter={'float_kind': float_formatter}) # encore utile ?

    stw.cell(row=k + 2, column=8).value = proba[0, 0]
    stw.cell(row=k + 2, column=9).value = proba[0, 1]

    coteHome = round(Decimal(indiceSecuriteHome * 1 / proba[0, 1]), 3)
    coteAway = round(Decimal(indiceSecuriteAway * 1 / proba[0, 0]), 3)

    stw.cell(row=k + 2, column=10).value = coteAway
    stw.cell(row=k + 2, column=11).value = coteHome

    print(teamAway, "@", teamHome, stats_final, proba, coteAway, coteHome)

    k += 1

today = str(date.today())
excel_file_name = "NBA_report_" + today + ".xlsx"
workbook.save(excel_file_name)